import csv
import json
import math
import subprocess
import sys
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from broiler_model.assumptions import Assumptions
from broiler_model.exporter import generate_excel_workbook
from broiler_model.model import generate_model_outputs
from broiler_model.production import (
    build_revenue_schedules,
    compute_cycles,
    annual_summary,
    summarise_revenue_totals,
)
from broiler_model.financing import discounted_cash_flow, build_financial_statements
from broiler_model.analytics import AnalyticsPlan, compute_advanced_analytics


class GenerateModelOutputsTests(unittest.TestCase):
    def test_generate_outputs_contains_expected_sections(self) -> None:
        assumptions = Assumptions()
        results = generate_model_outputs(assumptions)

        self.assertIs(results["assumptions"], assumptions)
        self.assertIn("valuation", results)
        valuation = results["valuation"]
        self.assertIn("npv", valuation)
        self.assertTrue(math.isfinite(valuation["npv"]))
        self.assertGreater(len(results["cashflows"]), 0)
        self.assertIn("advanced_analytics", results)
        self.assertIn("detail_schedule_outputs", results)

    def test_detailed_schedules_roll_up_into_model_assumptions(self) -> None:
        assumptions = Assumptions()
        detail_schedules = {
            "equipment_capex": [
                {
                    "Item": "Feed lines",
                    "Opening amount": 120000.0,
                    "New additions": 30000.0,
                    "Depreciation rate": 0.10,
                }
            ],
            "housing_capex": [
                {
                    "Item": "Broiler house A",
                    "Opening amount": 400000.0,
                    "New additions": 50000.0,
                    "Depreciation rate": 0.05,
                }
            ],
            "labor": [
                {
                    "Role": "Supervisors",
                    "Headcount": 2,
                    "Cost per head per cycle": 1800.0,
                },
                {
                    "Role": "Handlers",
                    "Headcount": 4,
                    "Cost per head per cycle": 950.0,
                },
            ],
            "maintenance": [
                {
                    "Item": "Preventive maintenance",
                    "Units": 1,
                    "Unit cost per cycle": 1750.0,
                }
            ],
            "management_fee": [
                {
                    "Item": "Farm manager",
                    "Units": 1,
                    "Unit cost per cycle": 2200.0,
                }
            ],
        }

        results = generate_model_outputs(assumptions, detail_schedules=detail_schedules)
        resolved = results["assumptions"]

        self.assertAlmostEqual(resolved.capex_equipment, 150000.0)
        self.assertAlmostEqual(resolved.capex_housing, 450000.0)
        self.assertAlmostEqual(resolved.labor_per_cycle, 7400.0)
        self.assertAlmostEqual(resolved.maintenance_per_cycle, 1750.0)
        self.assertAlmostEqual(resolved.management_fee_per_cycle, 2200.0)
        self.assertAlmostEqual(results["annual"].depreciation, 37500.0)

    def test_multi_loan_debt_schedule_supports_multiple_facilities(self) -> None:
        assumptions = Assumptions(capex_housing=500000.0, capex_equipment=250000.0)
        detail_schedules = {
            "debt_facilities": [
                {
                    "Loan name": "Senior term loan",
                    "Facility type": "Bank debt",
                    "Principal": 300000.0,
                    "Interest rate": 0.08,
                    "Term (years)": 5,
                    "Grace period (years)": 0,
                    "Start year": 1,
                    "Repayment type": "annuity",
                },
                {
                    "Loan name": "Equipment finance",
                    "Facility type": "Lease",
                    "Principal": 100000.0,
                    "Interest rate": 0.10,
                    "Term (years)": 4,
                    "Grace period (years)": 1,
                    "Start year": 2,
                    "Repayment type": "straight_line",
                },
            ]
        }

        results = generate_model_outputs(assumptions, detail_schedules=detail_schedules)
        resolved = results["assumptions"]
        debt_schedule = results["financial_statements"]["loan_schedule"]

        self.assertAlmostEqual(resolved.debt_ratio, 400000.0 / 750000.0, places=6)
        self.assertTrue(debt_schedule)
        self.assertGreaterEqual(len(debt_schedule), 4)
        self.assertGreater(debt_schedule[0]["payment"], 0.0)
        self.assertIn("calendar_year", debt_schedule[0])


class AdvancedAnalyticsTests(unittest.TestCase):
    def test_compute_advanced_analytics_returns_metrics(self) -> None:
        assumptions = Assumptions()
        cycles = compute_cycles(assumptions)
        annual = annual_summary(assumptions, cycles)
        cashflows, loan_schedule = discounted_cash_flow(assumptions, annual)
        revenue_schedules = build_revenue_schedules(assumptions, cycles)
        revenue_summary = summarise_revenue_totals(
            revenue_schedules,
            assumptions.cycles_per_year,
            assumptions.production_horizon_years,
            assumptions.production_start_year,
        )
        financials = build_financial_statements(assumptions, cashflows, loan_schedule)

        analytics = compute_advanced_analytics(
            assumptions,
            cashflows,
            financials["income_statement"],
            financials["balance_sheet"],
            revenue_summary,
            revenue_schedules,
            annual,
        )

        metrics = analytics["metrics"]
        metric_names = {entry["metric"] for entry in metrics}
        self.assertIn("Base case NPV", metric_names)
        self.assertIn("Average DSCR", metric_names)

        monte_carlo = analytics["monte_carlo"]
        self.assertIn("summary", monte_carlo)
        self.assertIn("settings", monte_carlo)
        self.assertEqual(
            monte_carlo["summary"].get("iterations"),
            monte_carlo["settings"].get("iterations"),
        )
        self.assertIn("distributions", monte_carlo["settings"])

        custom_definitions = analytics.get("custom_simulation_definitions", [])
        self.assertTrue(custom_definitions)

        custom_payload = analytics.get("custom_simulations", {})
        self.assertIn("delta_summary", custom_payload)

        break_even_rows = analytics.get("break_even", [])
        if break_even_rows:
            self.assertIn("Direct cost", break_even_rows[0])

    def test_summary_plan_skips_heavy_sections(self) -> None:
        assumptions = Assumptions()
        cycles = compute_cycles(assumptions)
        annual = annual_summary(assumptions, cycles)
        cashflows, loan_schedule = discounted_cash_flow(assumptions, annual)
        revenue_schedules = build_revenue_schedules(assumptions, cycles)
        revenue_summary = summarise_revenue_totals(
            revenue_schedules,
            assumptions.cycles_per_year,
            assumptions.production_horizon_years,
            assumptions.production_start_year,
        )
        financials = build_financial_statements(assumptions, cashflows, loan_schedule)

        analytics = compute_advanced_analytics(
            assumptions,
            cashflows,
            financials["income_statement"],
            financials["balance_sheet"],
            revenue_summary,
            revenue_schedules,
            annual,
            plan=AnalyticsPlan.summary(),
        )

        self.assertEqual(analytics["monte_carlo"]["summary"].get("iterations"), 0)
        self.assertFalse(analytics["custom_simulations"].get("results"))
        self.assertFalse(analytics["scenario_planning"])


class CLIBaselineRegressionTests(unittest.TestCase):
    def test_cli_outputs_within_expected_ranges(self) -> None:
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory() as tmpdir:
            out_dir = Path(tmpdir)
            cmd = [
                sys.executable,
                "deployable_financial_model.py",
                "--out",
                str(out_dir),
                "--formats",
                "csv",
                "json",
            ]
            subprocess.run(cmd, check=True, cwd=repo_root)

            valuation_path = out_dir / "valuation.json"
            self.assertTrue(valuation_path.exists())
            with valuation_path.open() as fh:
                valuation = json.load(fh)

            self.assertAlmostEqual(valuation["npv"], -810252.04, delta=1_500)
            self.assertAlmostEqual(valuation["irr"], -0.1684, delta=0.01)

            dscr_path = out_dir / "dscr_summary.csv"
            self.assertTrue(dscr_path.exists())
            with dscr_path.open() as fh:
                reader = csv.DictReader(fh)
                first_row = next(reader)

            self.assertAlmostEqual(float(first_row["dscr"]), 0.8973, delta=0.02)


class WorkbookExportTests(unittest.TestCase):
    def test_generate_excel_workbook_contains_redesigned_sheets(self) -> None:
        assumptions = Assumptions()
        results = generate_model_outputs(assumptions, analytics_plan=AnalyticsPlan.summary())

        workbook_bytes = generate_excel_workbook(
            assumptions,
            results,
            "Baseline",
            ai_settings={"provider": "openai", "model": "gpt-4o-mini"},
        )

        workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
        self.assertIn("Overview", workbook.sheetnames)
        self.assertIn("Revenue Dashboard", workbook.sheetnames)
        self.assertIn("Assumptions", workbook.sheetnames)
        self.assertIn("Equipment Capex Detail", workbook.sheetnames)
        self.assertIn("Debt Facilities", workbook.sheetnames)
        self.assertIn("Production Cycles", workbook.sheetnames)
        self.assertEqual(workbook["Overview"]["B2"].value, "Broiler Chicken Financial Model  |  Baseline")
        self.assertEqual(workbook["Revenue Dashboard"]["B2"].value, "Revenue Mix Dashboard")


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
